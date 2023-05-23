"""
ThousandEyes Tests Information

This script analyzes the tests for each account group in ThousandEyes
nd provides details about each test, including the name, test ID, test type,
protocol, probe method, and path trace settings.

Usage:
    python test_info_excel.py

Author: Miguel Hernandez
Date: May 22, 2023
"""
from getpass import getpass
from pprint import pprint
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook
import requests


def get_accounts(username, token):
    """
    Retrieves account information from the ThousandEyes API and returns a
    dictionary mapping account IDs to account names.

    Args:
        username (str): The username for authentication.
        token (str): The basic authentication token.

    Returns:
        dict: A dictionary mapping account IDs (aid) to account names (Account Group Name).

    Raises:
        requests.exceptions.RequestException: If there was an error making the API request.
    """
    headers = {"content-type": "application/json"}

    url = "https://api.thousandeyes.com/v6/account-groups.json"

    try:
        tests_api = requests.get(
            url, headers=headers, auth=(username, token), timeout=5
        )
    except requests.exceptions.Timeout:
        print("Get Accounts Exception: Request timed out to getting accounts")
    except requests.exceptions.RequestException as error:
        print("Get Accounts Exception:", str(error))
    response = tests_api.json()
    account_groups = response.get("accountGroups")
    accounts = {}
    for entry in account_groups:
        account_name = entry.get("accountGroupName")
        account_aid = entry.get("aid")
        accounts[account_aid] = account_name
    return accounts


def get_tests(username, token, aid):
    """
    Retrieves tests information from the ThousandEyes API and returns a
    dictionary with the details of each test.

    Args:
        username (str): The username for authentication.
        token (str): The basic authentication token.

    Returns:
        dict: A dictionary with test's information.

    Raises:
        requests.exceptions.RequestException: If there was an error making the API request.
    """
    headers = {"content-type": "application/json"}

    url = f"https://api.thousandeyes.com/v6/tests.json?aid={aid}"

    try:
        tests_api = requests.get(
            url, headers=headers, auth=(username, token), timeout=5
        )
    except requests.exceptions.Timeout:
        print("Get Tests Exception: Request timed out to getting accounts")
    except requests.exceptions.RequestException as error:
        print("Get Tests Exception:", str(error))
    response = tests_api.json()
    tests_configured = response.get("test")
    pprint(tests_configured)
    return tests_configured

# Provide the username and the basic token to retrieve the data from TE's API
USERNAME = input("Provide Username: ") or "miherna2@thousandeyes.com"
BASIC_TOKEN = getpass("Provide Basic Token: ") or "6wk0yr62ndkwhm6pta88prlsc5jmzlt8"

accounts_dict = get_accounts(USERNAME, BASIC_TOKEN)
print(accounts_dict)

# A new workbook is generated, the first sheet contains the Account Groups information.
workbook = Workbook()
workbook.remove(workbook.active)
sheet_accounts = workbook.create_sheet("ACCOUNTS")
for col in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
    sheet_accounts.column_dimensions[col].width = 25

sheet_accounts["A1"] = "AID"
sheet_accounts["B1"] = "ACCOUNT_NAME"
header_row_accounts = sheet_accounts[1]
header_font = Font(bold=True)
fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
for cell in header_row_accounts:
    cell.font = header_font
    cell.fill = fill

row = 2
for key, value in accounts_dict.items():
    sheet_accounts.cell(row=row, column=1).value = key
    sheet_accounts.cell(row=row, column=2).value = value
    row += 1

# A new sheet is created per account group identifier collect the tests
for account_identifier in accounts_dict:
    tests = get_tests(USERNAME, BASIC_TOKEN, account_identifier)
    sheet = workbook.create_sheet(str(account_identifier))
    for col in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
        sheet.column_dimensions[col].width = 25
    sheet.append(
        ["TEST_ID", "TEST_NAME", "TEST_TYPE", "PROTOCOL", "PROBE", "PATH_TRACE_MODE"]
    )
    header_row_aid = sheet[1]
    header_font = Font(bold=True)
    fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    for cell in header_row_aid:
        cell.font = header_font
        cell.fill = fill
    for test in tests:
        test_id = str(test.get("testId"))
        test_name = str(test.get("testName"))
        protocol = str(test.get("protocol"))
        probe = str(test.get("probeMode"))
        test_type = str(test.get("type"))
        path_trace = str(test.get("pathTraceMode"))
        sheet.append([test_id, test_name, test_type, protocol, probe, path_trace])
    workbook.save("output.xlsx")
